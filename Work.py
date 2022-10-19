import openpyxl
import datetime
wb=openpyxl.load_workbook("C:\\Users\\Lopez\\OneDrive\\WIP\\Test Excel.xlsx")
sh1=wb['Sheet1']
sh2=wb['Daily Amazon']
rows=sh1.max_row
rows2=sh2.max_row
print(rows2)

def ensure():
 global yesorno
 yesorno=input(str("Are you sure.(Y/N)\n"))
 while yesorno!='y' and yesorno!='n':
  print("Incorrect Input")
  yesorno=input("Are you sure\n")

def sure(function):
 function()
 ensure()   
 while yesorno=='n':
  function()
  ensure()

def sure2(product):
 addition=input(product)
 ensure()   
 while yesorno=='n':
  addition=input(product)
  ensure()
 return addition


#Offers options to choose what the user would like to do
def which():
    global guide
    guide=input(str("Which action would you like to perform\n1. Input New Item to Inventory\n2. Add Cost\n3. Profit\n4. Payout Per Person\n5. Check Total Cost\n"))
#Asks the user to double check if they made the right decision
restart='yes'
while restart=='yes':

sure(which)

#Matches the users choice to the function 
match guide:
  #------------------------------------------------------------------------------------------------------------
 case "1": 
  #Takes the user input of the item,cost and quantity
  #newitem=input("Input New item\n")
  newitem=sure2("Input New item\n")
  newcost=sure2("Input cost of Items\n")
  quantity=sure2('Input the quantity of Items\n')
  #For the amount of values between 1 and the number of rows, execute the blocks within
  #the for loop while increasing 'i'
  for i in range (1,rows+2):
    #if the cell check is not empty and has the same value of item string already inputted
    #update the quantity of items by the amount put in by the user and then break the for statement
    #if not an exisiting item, the block will skip the if statement 
    if sh1.cell(i,1).value !=None and sh1.cell(i,1).value==newitem:
        yesorno=input("This item has already been inputted\n")
        upquan=int(sh1.cell(i,3).value)
        upquan+=quantity
        sh1.cell(i,3).value=upquan 
        break 
    #if the cell value is equal to none aka the cell is empty the elif block while activate, otherwise it will move on
    #if the cell value is equal to none, the code will update the cell values with new items
    elif sh1.cell(i,1).value ==None:
     sh1.cell(i,1).value=newitem
     sh1.cell(i,2).value=newcost
     sh1.cell(i,3).value=quantity
#----------------------------------------------------------------------------------------------------------
 case "3": 
  
  sale=sure2(input('What is the name of the sold product\n'))
  salerev=sure2(input(str('What the revenue of the sale\n')))
  salefee=sure2(input(str('What is the fee of the sale\n')))
  #increase the value of i so long as it is inbetween 1 and the number of row +1
  #also only runs the for loops for the number of time between the 1 and row+1
  for i in range (1,rows+1):
   if sh1.cell(i,1).value==sale:
    print("A product with the same name has been found",i)
    for j in range (1,rows2+2):
      print("cell ",j," is empty")
      if sh2.cell(j,1).value==None: 
       print("row ",j,"column 1 is empty")
       sh2.cell(j,2).value=sale
       #sets item cost on sheet 2
       sh2.cell(j,5).value=sh1.cell(i,2).value 
       #sets
       sh2.cell(j,3).value=salerev  
       sh2.cell(j,4).value=salefee
       sh2.cell(j,1).value= datetime.date.today()
       sh2.cell(j,6).value=float(salerev)-float(salefee)-float(sh2.cell(j,5).value)
       break  
  

wb.save("Test Excel.xlsx")



    


"""
 case '2':
  costname=input("Input Cost Name\n")
  cost= input("Input Cost\n")
 case '3':
  print("The total profit is")
 case '4':
  print("The Payout Per Person is")
 case '5':
  print("The total cost is ")
"""