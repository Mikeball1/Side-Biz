from tkinter import W
import openpyxl 
#Load excel sheet and assign its type to a variable
wb=openpyxl.load_workbook("C:\\Users\\Lopez\\OneDrive\\WIP\\Actual Amazon.xlsx")
#Prints object type
print(type(wb))
#Store the number of sheets in a variabel
sheets=wb.sheetnames
#Prints the number of sheets
print(sheets)
#Prints the sheets the is currently open
print(wb.active.title)
#-----------------------------------------
#Assigns the sheet to a variable
sh1=wb['Sheet1']
#Prints the sheet type
print(type(sh1))
#returns the data within sheet 
data=sh1['E3'].value 
print(data)
#-----------------------------------------
#Conslidates all the work above
data2=wb['Sheet1']['E3'].value
print(data2)
#------------------------------------------
#Prints the max rows and columns you have in the sheet
rows=sh1.max_row
print(rows)
col=sh1.max_column
print(sh1.max_column)
#------------------------------------------
#prints all the information in the first row
for i in range (1,rows+1):
 #for j in range (1,col):
  print(sh1.cell(i,1).value )
#-------------------------------------------
#prints all the info in the first columns
for j in range(1,col+1):
 print (sh1.cell(1,j).value)
#-------------------------------------------
sh1['D2'].value='wow'
if sh1['D2'].value ==None: 
 print ("This cell is empty")
else:
  print ("This cell is not empty") 
#saves file
wb.save("Test Excel.xlsx")