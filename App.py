import pip
from tkinter import *
import openpyxl
from datetime import datetime
import random 
import webbrowser
#Opens main excel sheet
wb=openpyxl.load_workbook("Test Excel.xlsx")
sh1=wb['Sheet1']
sh2=wb['Daily Amazon']
sh3=wb['Pay Period']
rows=sh1.max_row
rows2=sh2.max_row
rows3=sh3.max_row

class NewItem():
 def __init__(self,Name,Cost,Quantity):
    self.Name=Name
    self.Cost=Cost
    self.Quantity=Quantity
#----------------------------------------------------------------------------------------
#Input new item to inventory or update inventory
def input_new_item():
    input_window = Toplevel()
    names=Label(input_window,text="Item Name")
    costs=Label(input_window,text="Item Costs")
    amounts=Label(input_window,text="Item Quantity")
    names.grid(row=1,column=0)
    costs.grid(row=2,column=0)
    amounts.grid(row=3,column=0)
    input_window.title("Input New Item")
    inputs=Label(input_window,text="Input the name of the new item")
    inputs.grid(row=0,column=1)
    NewitemName=Entry(input_window,width=50)
    NewitemName.grid(row=1,column=1)
    NewitemAmount=Entry(input_window,width=50)
    NewitemAmount.grid(row=2,column=1)
    NewitemCost=Entry(input_window,width=50)
    NewitemCost.grid(row=3,column=1)    
    inventorys=Label(input_window,text="Current Inventory")
    inventorys.grid(row=0,column=2)
    for name in range(2,rows+1):
       inventory=Label(input_window,text=sh1.cell(name,1).value)
       inventory.grid(row=name-1,column=2)
    def show_input():
     global Item
     Item=NewItem(NewitemName.get(),NewitemCost.get(),NewitemAmount.get())
     itemDes=Label(input_window,text="The name of your inputted item is "+Item.Name+"\nThe cost of your inputted item is "+Item.Cost+"\nThe quantity of the item is "+ Item.Quantity)
     itemDes.grid(row=6,column=1)

     if type(Item.Name)==str:
      for i in range (1,rows+2):
       if sh1.cell(i,1).value !=None and sh1.cell(i,1).value==Item.Name:
          yesorno=Label(input_window,text="This item has already been inputted\n")
          yesorno.grid(row=7,column=1)
          upquan=int(sh1.cell(i,3).value)
          upquan+=int(Item.Quantity)
          sh1.cell(i,3).value=upquan 
          wb.save("Test Excel.xlsx")
          break 
       elif sh1.cell(i,1).value ==None:
        sh1.cell(i,1).value=(Item.Name)
        sh1.cell(i,2).value=int(Item.Cost)
        sh1.cell(i,3).value=int(Item.Quantity)
        wb.save("Test Excel.xlsx")
        break
    Characteristics=Button(input_window,text="Press to lock in values",command=show_input)
    Characteristics.grid(row=4,column=1)
#----------------------------------------------------------------------------------------
def enter_sale():
    inventory=[]
    payout_window = Toplevel()
    payout_window.title("Enter_sale")
    profit=Label(payout_window,text="Enter Sales")
    profit.grid(row=0,column=1)
    salenames=Label(payout_window,text="Item Name")
    salesrev=Label(payout_window,text="Item Revenue")
    salesfee=Label(payout_window,text="Item Fee")
    salenames.grid(row=1,column=0)
    salesfee.grid(row=2,column=0)
    salesrev.grid(row=3,column=0)
    for name in range(2,rows+1):
       inventory.append(sh1.cell(name,1).value)
    chosen=StringVar()
    chosen.set("Click to show items available")
    salename=OptionMenu(payout_window,chosen,*inventory)
    salename.grid(row=1,column=1)
    salerev=Entry(payout_window,width=20)
    salerev.grid(row=2,column=1)
    salefee=Entry(payout_window,width=20)
    salefee.grid(row=3,column=1)
    lastsales=Label(payout_window,text="The last three sales are")
    lastsales.grid(row=0,column=2)
    #displays the last three sales
    x=0
    for e in range(rows2,rows2-3,-1):
      x+=1
      date=(sh2.cell(e,1).value)
      date_only=date.strftime("%m/%d/%Y")
      text = date_only + "-" + (sh2.cell(e,2).value) + "-" + str(sh2.cell(e,3).value) + "-" + str(sh2.cell(e,4).value) + "\n"
      prevsale=Label(payout_window, text=text)
      prevsale.grid(row=x,column=2)
    # Add space for the coder to input their code here
    def lockin():
        
        Sales=NewItem(chosen.get(),salerev.get(),salefee.get())
        for i in range (1,rows+1):
         if sh1.cell(i,1).value==Sales.Name:
           print("A product with the same name has been found")
           sh2.cell(rows2+1,2).value=Sales.Name
           sh2.cell(rows2+1,3).value=float(Sales.Cost)
           sh2.cell(rows2+1,4).value=float(Sales.Quantity)
           sh2.cell(rows2+1,1).value=date.today()
           #Sets item cost on sheet 2
           sh2.cell(rows2+1,5).value=sh1.cell(i,2).value
           sh2.cell(rows2+1,6).value=float(Sales.Cost)-float(Sales.Quantity)-float(sh2.cell(rows2+1,5).value)
            #once the program has found an item from the first sheet, itll input all the values into the 2nd sheet
            #then it will proceed to count from bottom to top and check who got the last sale for said item
            #and based on that, determine who gets the next sale
           for k in range(rows2,1,-1):
             if sh2.cell(k,2).value==Sales.Name:
              if sh2.cell(k,7).value==1:
               print('This sale goes to Matthew')
               sh2.cell(rows2+1,8).value=1
               wb.save("Test Excel.xlsx")
               break
              if sh2.cell(k,8).value==1:
               print('This sale goes to Michael')
               wb.save("Test Excel.xlsx")
               sh2.cell(rows2+1,7).value=1
               break
                #this if statement determines if the last time the info was inputted, if the info had who the sale belonged to
             if k==2: #sh2.cell(k,7).value!=1 and sh2.cell(k,8).value!=1:
              rando=random.randint(1,2)
              if rando==2: 
                print('This sale goes to Michael as determined rng')
                sh2.cell(rows2+1,7).value=1
                wb.save("Test Excel.xlsx")
                break
              if rando==1:
               print('This sale goes to Matthew as determined by rng')
               sh2.cell(rows2+1,8).value=1
               wb.save("Test Excel.xlsx")
               break
           break 
        
    sales=Button(payout_window,text="Press to lock in values",command=lockin)
    sales.grid(row=4,column=1)
    wb.save("Test Excel.xlsx")

 
def payout():
    payout_window = Toplevel()
    payout_window.title("Payout")
    # Add space for the coder to input their code here
    pass

def item_sales_per_person():
    payout_window = Toplevel()
    payout_window.title("item_sales_per_person")
    # Add space for the coder to input their code here
    pass
def sales_per_person():
    payout_window = Toplevel()
    payout_window.title("sales_per_person")
    # Add space for the coder to input their code here
    pass
def end_program():
    payout_window = Toplevel()
    payout_window.title("end_program")
    # Add space for the coder to input their code here
    pass
def add_shipment():
    payout_window = Toplevel()
    payout_window.title("add_shipment")
    # Add space for the coder to input their code here
    pass
def refund():
    payout_window = Toplevel()
    payout_window.title("refund")
    # Add space for the coder to input their code here
    pass
def other_fees():
    payout_window = Toplevel()
    payout_window.title("other_fees")
    # Add space for the coder to input their code here
    pass
def current_sales():
    payout_window = Toplevel()
    payout_window.title("current_sales")
    # Add space for the coder to input their code here
    pass
def current_profits():
    payout_window = Toplevel()
    payout_window.title("current_profits")
    # Add space for the coder to input their code here
    pass
#sure function is used for running other functions with the yes or no prompt

# def ensure():
#  global yesorno
#  yesorno=input(str("Are you sure.(Y/N)\n"))
#  while yesorno!='y' and yesorno!='n':
#   print("Incorrect Input")
#   yesorno=input("Are you sure\n")
# #sure function is used for running input with the yes or no prompt
# def sure(function):
#  function()
#  ensure()   
#  while yesorno=='n':
#   function()
#   ensure()
def case1():
    pass