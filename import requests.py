import pip
import openpyxl
import datetime
import random 
import webbrowser
webbrowser.open('https://sellercentral.amazon.com/payments/event/view?resultsPerPage=10&pageNumber=1')

def ensure():
 global yesorno
 yesorno=input(str("Are you sure.(Y/N)\n"))
 while yesorno!='y' and yesorno!='n':
  print("Incorrect Input")
  yesorno=input("Are you sure\n")
#sure function is used for running other functions with the yes or no prompt
def sure(function):
 function()
 ensure()   
 while yesorno=='n':
  function()
  ensure()
#sure2 function is used for running input with the yes or no prompt
def sure2(product):
 addition=input(product)
 ensure()   
 while yesorno=='n':
  addition=input(product)
  ensure()
 return addition
def case3():
    global salename 
    global salerev
    global salefee
    salename=(sure2('What is the name of the sold product.\n')).strip()
    salerev=float(sure2('What the revenue of the sale\n'))
    salefee=float(sure2('What is the fee of the sale\n'))
    #increase the value of i so long as it is inbetween 1 and the number of row +1
    #also only runs the for loops for the number of time between the 1 and row+1
    for i in range (1,rows+1):
     if sh1.cell(i,1).value==salename:
      print("A product with the same name has been found")
      sh2.cell(rows2+1,2).value=salename
      sh2.cell(rows2+1,3).value=salerev  
      sh2.cell(rows2+1,4).value=salefee
      sh2.cell(rows2+1,1).value=datetime.date.today()
      #sets item cost on sheet 2
      sh2.cell(rows2+1,5).value=sh1.cell(i,2).value
      sh2.cell(rows2+1,6).value=float(salerev)-float(salefee)-float(sh2.cell(rows2+1,5).value)
      #once the program has found an item from the first sheet, itll input all the values into the 2nd sheet
      #then it will proceed to count from bottom to top and check who got the last sale for said item
      #and based on that, determine who gets the next sale
      for k in range(rows2,1,-1):
       if sh2.cell(k,2).value==salename:
        if sh2.cell(k,7).value==1:
          print('This sale goes to Matthew')
          sh2.cell(rows2+1,8).value=1
          break
        if sh2.cell(k,8).value==1:
          print('This sale goes to Michael')
          sh2.cell(rows2+1,7).value=1
          break
        #this if statement determines if the last time the info was inputted, if the info had who the sale belonged to
       if k==2: #sh2.cell(k,7).value!=1 and sh2.cell(k,8).value!=1:
         rando=random.randint(1,2)
         if rando==2: 
          print('This sale goes to Michael as determined rng')
          sh2.cell(rows2+1,7).value=1
          break
         if rando==1:
          print('This sale goes to Matthew as determined by rng')
          sh2.cell(rows2+1,8).value=1
          break
      break 
      #this if goes through all the sales and if it reaches the bottom of the list and the 
      #sheet itself has never had that product inputted, it will assign the first sale at random
'''
       if k==1:
        print("rng calculation")
        rando=random.randint(1,2)
        if rando==2:
          sh2.cell(rows2+1,7).value=1
        else:
          sh2.cell(rows2+1,8).value=1
        break  
'''

#Offers options to choose what the user would like to do
def which():
    choice=['1','2','3','4','5','6','7','8','9','10','11']
    global guide
    guide=input(str("Which action would you like to perform\n1. Input New Item to Inventory\n2. Payout\n3. Enter Sale \n4. Item sales Per Person\n5. Sales Per Person\n6. End Program\n7. Add Shipment\n8. Refund\n9. Other Fees\n10. Current Sales\n11. Current Profits\n"))
    while guide not in choice:
      print("Please pick one of the options")
      guide=input(str("Which action would you like to perform\n1. Input New Item to Inventory\n2. Payout\n3. Enter Sale \n4. Item sales Per Person\n5. Sales Per Person\n6. End Program\n7. Add Shipment\n8. Refund\n9. Other Fees\n10. Current Sales\nCurrent Profits\n"))

#Asks the user to double check if they made the right decision
restart='y'
while restart=='y':
  wb=openpyxl.load_workbook("Test Excel.xlsx")
  sh1=wb['Sheet1']
  sh2=wb['Daily Amazon']
  sh3=wb['Pay Period']
  rows=sh1.max_row
  rows2=sh2.max_row
  rows3=sh3.max_row

  sure(which)

  #Matches the users choice to the function 
  match guide:
    #------------------------------------------------------------------------------------------------------------
   case "1": 
    #Takes the user input of the item,cost and quantity
    #newitem=input("Input New item\n")
    print("-------------------------")
    for name in range(1,rows+1):
     print(sh1.cell(name,1).value)
    print(" -------------------------")
    newitem=(sure2("Input New item\n")).strip()
    newcost=float(sure2("Input cost of Items\n"))
    quantity=int(sure2('Input the quantity of Items\n'))
    #For the amount of values between 1 and the number of rows, execute the blocks within
    #the for loop while increasing 'i'
    for i in range (1,rows+2):
      #if the cell check is not empty and has the same value of item string already inputted
      #update the quantity of items by the amount put in by the user and then break the for statement
      #if not an exisiting item, the block will skip the if statement 
      if sh1.cell(i,1).value !=None and sh1.cell(i,1).value==newitem:
          yesorno=input("This item has already been inputted\n")
          upquan=int(sh1.cell(i,3).value)
          upquan+=quantity
          sh1.cell(i,3).value=upquan 
          break 
      #if the cell value is equal to none aka the cell is empty the elif block while activate, otherwise it will move on
      #if the cell value is equal to none, the code will update the cell values with new items
      elif sh1.cell(i,1).value ==None:
       sh1.cell(i,1).value=newitem
       sh1.cell(i,2).value=newcost
       sh1.cell(i,3).value=quantity
       break

  #----------------------------------------------------------------------------------------------------------
   case "2":
    reserve=float(input("Input Reserve fees: "))
    totalsales=0
    fees=0
    for p in range(rows2,0,-1):
     if type(sh2.cell(p,3).value)==str:
      for m in range(rows2,p,-1):
       totalsales+=sh2.cell(m,3).value
      for fee in range(rows2,p,-1):
        if type(sh2.cell(fee,4).value)==float:
         fees+=(sh2.cell(fee,4).value)
      break
    global mikeprof
    mikeprof=0
    mattprof=0
    mikerev=0
    mattrev=0
    mikefee=0
    mattfee=0
    feeamount=0
    for p in range(rows2,0,-1):
      if type(sh2.cell(p,3).value)==str:
        for mfee in range(rows2,p,-1):
          if sh2.cell(mfee,8).value==1:
            mattfee+=sh2.cell(mfee,4).value
          if sh2.cell(mfee,7).value==1:
            mikefee+=sh2.cell(mfee,4).value
          if float(sh2.cell(mfee,3).value)<0:
            feeamount+=sh2.cell(mfee,3).value
        for m in range(rows2,p,-1):
         if sh2.cell(m,7).value==1:
         #if the 8th column of row m contains a value of 1 then add the name of that item to the list
          mikeprof+=sh2.cell(m,6).value
          mikerev+=sh2.cell(m,3).value
         if sh2.cell(m,8).value==1:
         #if the 8th column of row m contains a value of 1 then add the name of that item to the list
          mattprof+=sh2.cell(m,6).value
          mattrev+=sh2.cell(m,3).value
      break
    randomfee=float(sure2('Input any random fees or something\n'))
    print('Total Fees:', round(fees,2))
    print("Total Sales:", round(totalsales,2))
    print("Total Revenue:",round(totalsales-fees+randomfee+sh3.cell(rows3,3).value, 2))
    print("Total Payout:",round(totalsales-fees-reserve+randomfee+sh3.cell(rows3,3).value, 2))
    print("Matthew gets",round((feeamount+randomfee-reserve+sh3.cell(rows3,3).value)/2-mattfee+mattrev, 2),"\nMichael gets",round((feeamount+randomfee-reserve+sh3.cell(rows3,3).value)/2-mikefee+mikerev, 2))
    print("On the next payout,",round(reserve/2, 2),"will go to Matthew and Michael") 
    sh3.cell(rows3+1,3).value=reserve
    sh3.cell(rows3+1,5).value=round((feeamount+randomfee-reserve+sh3.cell(rows3,3).value)/2-mattfee+mattrev, 2)
    sh3.cell(rows3+1,4).value=round((feeamount+randomfee-reserve+sh3.cell(rows3,3).value)/2-mikefee+mikerev, 2)
    sh3.cell(rows3+1,2).value=sh3.cell(rows3+1,5).value+sh3.cell(rows3+1,4).value
    sh3.cell(rows3+1,1).value=datetime.date.today() 

  #----------------------------------------------------------------------------------------------------------
   case "3": 
    print("The Last 3 sales are:")
    for e in range(rows2,rows2-3,-1):
      print(sh2.cell(e,1).value,"-",sh2.cell(e,2).value,"-",sh2.cell(e,3).value,"-",sh2.cell(e,4).value,"\n")
    case3()
    while sh2.cell(rows2+1,2).value==None:
     print("No item with that name has been found, please try again")
     case3()
  #-----------------------------------------------------------------------------------------------------------
   case "4":
    itematt=[] 
    itemattrev=[]
    itemike=[]
    itemikerev=[]
    mikecount=[]
    mattcount=[]
    mikearr=[]
    mattarr=[]
    numitematt=[]
    numitemike=[]
    item1matt=[]
    item1mike=[]
    finarmat=[]
    finarmike=[]
    #payout=sure2("Input payout\n")
    for p in range(rows2,0,-1):
      #starts a bottom to top counter of the rows in the sheet
      if type(sh2.cell(p,3).value)==str:
       #checks to see which cell is not an integer, indicating that was the last pay period
       print("the last pay period was found on row",p)
       for m in range(rows2,p,-1):
        #once  it is determined where the last period ended, a new for loop within that range will 
        #start counting down from bottom to top
        if sh2.cell(m,7).value==1:
         #if the 8th column of row m contains a value of 1 then add the name of that item to the list
         itemike.append(sh2.cell(m,2).value)
        if sh2.cell(m,8).value==1:
         #if the 8th column of row m contains a value of 1 then add the name of that item to the list
          itematt.append(sh2.cell(m,2).value)
       break
    for n in itemike:
      if n not in item1mike:
        item1mike.append(n)
    for b in item1mike:
      numitemike.append(itemike.count(b))
    for c in item1mike:
      v=item1mike.index(c)
      r=numitemike[v]
      tup2=(c,r)
      finarmike.append(tup2)
#------------------------------------------------------------
    for j in itematt:
     if j not in item1matt:
      item1matt.append(j)
    for l in item1matt:
      numitematt.append(itematt.count(l))
    for o in item1matt:
      t=item1matt.index(o)
      q=numitematt[t]
      tup=(o,q)
      finarmat.append(tup)
    print ("Matthews product sales are:\n",finarmat)
    print("Michaels product sales are:\n",finarmike)
     
   case "5":
    mikeprof=0
    mattprof=0
    mikerev=0
    mattrev=0
    for p in range(rows2,0,-1):
      if type(sh2.cell(p,3).value)==str:
        for m in range(rows2,p,-1):
         if sh2.cell(m,7).value==1:
         #if the 8th column of row m contains a value of 1 then add the name of that item to the list
          mikeprof+=sh2.cell(m,6).value
          mikerev+=sh2.cell(m,3).value
         if sh2.cell(m,8).value==1:
         #if the 8th column of row m contains a value of 1 then add the name of that item to the list
          mattprof+=sh2.cell(m,6).value
          mattrev+=sh2.cell(m,3).value
        
    print("----------------------------\nMichaels profit:", mikeprof,"\nMichaels revenue:",mikerev)
    print("----------------------------\nMatthews profit:", mattprof,"\nMatthews revenue:",mattrev)

      # for m in range(rows2+1,p,-1):
         # print("checking row",m,"for each person")
       #  if sh2.cell(m,8).value==1:
        #  print("this one goes to matthew")
          
    #print ("matthew array",arr2)
    #print("michael array",arr)
      
   case '6':
    break 
   case '7':
    x=0
    y=0
    z=0
    tr=0
    shipnamelist=[]
    shipquanlist=[]
    shipcostlist=[]
    floatlist=[]
    combinelist=[]
    ques='y'
    while ques=='y':

     shipname=sure2("Input Item Name: ")
     shipquan=sure2("Input Quantity: ")
     shipitemcost=float(sure2(("How much did this product cost: ")))

     shipnamelist.append(shipname)
     shipquanlist.append(shipquan)
     shipcostlist.append(shipitemcost)
    
     ques=input('would you like to input more (y/n): ')
    for h in shipquanlist:
     floatlist.append(float(h))
    
    totalshipquan=sum(floatlist)
    
    shipfee=input("How much is the shipping fee: ")
    split=int(shipfee)/totalshipquan
    for v in shipcostlist:
      lop=shipcostlist.index(v)
      shipcostlist[lop]=float(v)+split
    for item in shipnamelist:
      x+=1
      itempos=shipnamelist.index(item)
      q=shipcostlist[itempos]
      tup=(item,q)
      combinelist.append(tup)
      x+=1
      print(combinelist)
      for thing in shipnamelist:
       sh1.cell(rows+1+tr,1).value=thing
       tr+=1
      for it in shipquanlist:
       sh1.cell(rows+1+y,3).value=it
       y+=1
      for what in shipcostlist:
       sh1.cell(rows+1+z,2).value=int(what)
       z+=1
   case '8':
    refundname=str(sure2("Which item was refunded\n"))
    refundamount=float(sure2('How much was the refund total. Input as positive number\n'))
    sh2.cell(rows2+1,1).value=datetime.date.today()
    sh2.cell(rows2+1,2).value=refundname + 'REF'
    sh2.cell(rows2+1,3).value=refundamount*-1
   case '9':
    randfeename=str(sure2("What is the name of the random fee\n"))
    randfee=float(sure2('How much was the random fee\n'))
    sh2.cell(rows2+1,1).value=datetime.date.today()
    sh2.cell(rows2+1,2).value=randfeename
    sh2.cell(rows2+1,3).value=randfee*-1
   case '10':
     salescurrent=0
     currsales2=0
     currfees2=0
     for p in range(rows2,0,-1):
      if type(sh2.cell(p,3).value)==str:
        for current in range(rows2,p,-1):
          if (sh2.cell(current,4).value)==None:
            sh2.cell(current,4).value=0
          currsales=float(sh2.cell(current,3).value)
          currfees=float(sh2.cell(current,4).value)
          currsales2+=float(sh2.cell(current,3).value)
          currfees2+=float(sh2.cell(current,4).value)
          salescurrent+=float(sh2.cell(current,3).value)-float(sh2.cell(current,4).value)
        print('The current balance should be ',salescurrent+sh3.cell(rows3,3).value)
        print("fee=",currfees2)
        print("sales=",currsales2)
        break
   case '11':
    currentprofits=0
    for p in range(rows2,0,-1):
        if type(sh2.cell(p,3).value)==str:
          for prof in range(rows2,p,-1):
            currentprofits+=sh2.cell(prof,6).value
          break
    print("The profits since the last pay period art",currentprofits)

  wb.save("Test Excel.xlsx")
  restart=input("Would you like to input more. (y/n)")


      


  
 