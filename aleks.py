from App import *
from tkinter import *
import pip
import openpyxl
import datetime
import random 
import webbrowser
# Define the functions that open new windows
tk = Tk()
tk.title("Inventory Management System")

#Opens main excel sheet
wb=openpyxl.load_workbook("Test Excel.xlsx")
sh1=wb['Sheet1']
sh2=wb['Daily Amazon']
sh3=wb['Pay Period']
rows=sh1.max_row
rows2=sh2.max_row
rows3=sh3.max_row

# Create the buttons
input_button = Button(tk, text="Input New Item", command=input_new_item)
payout_button = Button(tk, text="Payout", command=payout)
sale_button = Button(tk, text="Enter Sale", command=enter_sale)
item_sales_button = Button(tk, text="Item Sales Per Person", command=item_sales_per_person)
sales_button = Button(tk, text="Sales Per Person", command=sales_per_person)
end_button = Button(tk, text="End Program", command=end_program)
shipment_button = Button(tk, text="Add Shipment", command=add_shipment)
refund_button = Button(tk, text="Refund", command=refund)
fees_button = Button(tk, text="Other Fees", command=other_fees)
sales_button = Button(tk, text="Current Sales", command=current_sales)
profits_button = Button(tk, text="Current Profits", command=current_profits)

# Place the buttons on the window
input_button.pack()
payout_button.pack()
sale_button.pack()
item_sales_button.pack()
sales_button.pack()
end_button.pack()
shipment_button.pack()
refund_button.pack()
fees_button.pack()
sales_button.pack()
profits_button.pack()

# Run the main loop
tk.mainloop()
